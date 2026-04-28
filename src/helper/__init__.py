import pandas as pd
import base64, json, os
from .J2S import JSONtoSheets
from io import BytesIO
import openpyxl

def get_base_path():
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

file_dir = os.path.join(get_base_path(), 'data')
VALIDATION_CONFIG = JSONtoSheets().readconfig(os.path.join(file_dir, 'config.json'))
INTERMEDIATE_CONFIG = JSONtoSheets().readconfig(os.path.join(file_dir, 'intermediate_config.json'))

_global_config_path = os.path.join(file_dir, 'global_config.json')
with open(_global_config_path, 'r') as f:
    GLOBAL_CONFIG = json.load(f)

def get_files_data():
    variables = [sheet.Variable for sheet in VALIDATION_CONFIG]
    forecast_sheets = ["BOM_Forecast_Data_Future", "BOM_Forecast_Data_History"]
    other_sheets = [sheet.Variable for sheet in VALIDATION_CONFIG if sheet.Variable not in forecast_sheets]
    return {"variables": variables, "forecast_sheets": forecast_sheets, "other_sheets": other_sheets}

FILES = get_files_data()

def convert_df_to_base64(df):
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    return base64.b64encode(output.read()).decode('utf-8')

def base64_to_dataframe(base64_str):
    decoded_data = base64.b64decode(base64_str)
    data_io = BytesIO(decoded_data)
    df = pd.read_excel(data_io)
    return df

def add_data_to_excel_with_formatting(existing_file_path, dataframe):
    workbook = openpyxl.load_workbook(existing_file_path)
    sheet = workbook.active
    start_row = 2
    for c_idx, col_name in enumerate(dataframe.columns, start=1):
        cell = sheet.cell(row=1, column=c_idx)
        cell.value = col_name
    for r_idx, row in dataframe.iterrows():
        for c_idx, (col_name, value) in enumerate(row.items(), start=1):
            cell = sheet.cell(row=start_row + r_idx, column=c_idx)
            cell.value = value
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return base64.b64encode(output.getvalue()).decode('utf-8')

def update_global_config(new_config: dict):
    global GLOBAL_CONFIG
    GLOBAL_CONFIG.update(new_config)
